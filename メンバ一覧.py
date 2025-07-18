##########################################################################################################
#[状態]
#完成
#ロードボタンとセーブボタンを作ってもよい。作った場合はキャッシュデータであるmemberだけ更新する
#[機能]
#base_paramaterのシートの部分の編集
#参加者の決定、memberシートの更新
##########################################################################################################

import streamlit as st
import pandas as pd

#エクセル操作
import openpyxl

st.image("OpenChat.png")

wb=openpyxl.load_workbook("data.xlsx",data_only=True)
ws_member=wb["member"]
#############################
#メンバーシートの列定義
#############################
#MEMBER_ATTENDANCE_COLUMN=1
MEMBER_NAME_COLUMN=2
MEMBER_SEX_COLUMN=3
#MEMBER_LEVEL_COLUMN=4
MEMBER_DOUBLES_COLUMN=5
MEMBER_SINGLES_COLUMN=6
MEMBER_MIXED_COLUMN=7
#MEMBER_STATUS_COLUMN=8
MEMBER_POINT_COLUMN=9
MEMBER_NOT_COLUMN=10

##########################################################################################################
#画面の読み取り
##########################################################################################################
df = pd.DataFrame([{}])
df_cache = pd.DataFrame([{}])

for number in range(30):
    if ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value==None:
        break

    if number==0:
        df["名前"]=ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value
        df["性別"]=ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value
        df["ポイント"]=ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value
        df["回数"]=ws_member.cell(row=number+2,column=MEMBER_NOT_COLUMN).value
        if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
            df["ダブルス"]=True
        else:
            df["ダブルス"]=False
        if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
            df["シングルス"]=True
        else:
            df["シングルス"]=False
        if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
            df["ミックス"]=True
        else:
            df["ミックス"]=False
    else:
        df_cache["名前"]=ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value
        df_cache["性別"]=ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value
        df_cache["ポイント"]=ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value
        df_cache["回数"]=ws_member.cell(row=number+2,column=MEMBER_NOT_COLUMN).value
        if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
            df_cache["ダブルス"]=True
        else:
            df_cache["ダブルス"]=False
        if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
            df_cache["シングルス"]=True
        else:
            df_cache["シングルス"]=False
        if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
            df_cache["ミックス"]=True
        else:
            df_cache["ミックス"]=False

        df=pd.concat([df,df_cache],ignore_index=True)    

edited_df = st.data_editor(df,use_container_width=True,hide_index=True,height=600)

col1,col2=st.columns(2)

if col2.button("セーブ",use_container_width=True):

    for number1 in range(len(df)):
        for number2 in range(len(df)):
            #名前が一致したら上書き修正
            if edited_df.iat[number1,0]==ws_member.cell(row=number2+2,column=MEMBER_NAME_COLUMN).value:
                ws_member.cell(row=number2+2,column=MEMBER_NAME_COLUMN,value=edited_df.iat[number1,0])#名前
                ws_member.cell(row=number2+2,column=MEMBER_SEX_COLUMN,value=edited_df.iat[number1,1])#性別
                ws_member.cell(row=number2+2,column=MEMBER_POINT_COLUMN,value=edited_df.iat[number1,2])#ポイント
                ws_member.cell(row=number2+2,column=MEMBER_NOT_COLUMN,value=edited_df.iat[number1,3])#回数
                ws_member.cell(row=number2+2,column=MEMBER_DOUBLES_COLUMN,value=edited_df.iat[number1,4])#ダブルス
                ws_member.cell(row=number2+2,column=MEMBER_SINGLES_COLUMN,value=edited_df.iat[number1,5])#シングルス
                ws_member.cell(row=number2+2,column=MEMBER_MIXED_COLUMN,value=edited_df.iat[number1,6])#ミックス
                break

    wb.save("data.xlsx")



##########################################################################################################
#説明
##########################################################################################################
"[使い方]"
"基本操作の必要はないが、シングルとミックスなどの情報を修正したい場合はメンバリスト(member)を直接書き換えられる。"
"操作したらセーブボタンを押さないと反映されない。"

"[説明]"
"メンバーリスト(member)の一部の情報を修正できる。"
"「セーブ」ボタンを押すと画面の情報がメンバーリスト(member)に反映される。"
"変更されるのはあくまで当日のメンバとなるので、元データの更新はされない。"

