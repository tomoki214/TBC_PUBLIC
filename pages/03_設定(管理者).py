##########################################################################################################
#[状態]
#完成
#[機能]
#base_paramaterのシートの部分の編集
#参加者の決定、memberシートの更新
##########################################################################################################


import streamlit as st
import pandas as pd

#エクセル操作
import openpyxl

wb=openpyxl.load_workbook("data.xlsx",data_only=True)
ws_base_parameter=wb["base_parameter"]
ws_member=wb["member"]
ws_court=wb["court"]
ws_history=wb["history"]

#############################
#ベースパラメータシートの列定義
#############################
BASEPARAMETER_ATTENDANCE_COLUMN=1
BASEPARAMETER_NAME_COLUMN=2
BASEPARAMETER_SEX_COLUMN=3
BASEPARAMETER_LEVEL_COLUMN=4
BASEPARAMETER_DOUBLES_COLUMN=5
BASEPARAMETER_SINGLES_COLUMN=6
BASEPARAMETER_MIXED_COLUMN=7

#############################
#メンバーシートの列定義
#############################
MEMBER_ATTENDANCE_COLUMN=1
MEMBER_NAME_COLUMN=2
MEMBER_SEX_COLUMN=3
MEMBER_LEVEL_COLUMN=4
MEMBER_DOUBLES_COLUMN=5
MEMBER_SINGLES_COLUMN=6
MEMBER_MIXED_COLUMN=7
MEMBER_STATUS_COLUMN=8
MEMBER_POINT_COLUMN=9
MEMBER_NOT_COLUMN=10

##########################################################################################################
#画面の読み取り
##########################################################################################################
df = pd.DataFrame([{}])
df_cache = pd.DataFrame([{}])

for number in range(500):
    if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value==None:
        break
    if number==0:
        #参加不参加
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_ATTENDANCE_COLUMN).value=="〇":
            df["参加"]=True
        else:
            df["参加"]=False
        #名前
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SEX_COLUMN).value=="男":
            if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value.startswith("🔵")==False:
                df["名前"]="🔵"+ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
            else:
                df["名前"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
        else:
            if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value.startswith("🔴")==False:
                df["名前"]="🔴"+ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
            else:
                df["名前"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
        #性別
        df["性別"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SEX_COLUMN).value
        #レベル
        df["レベル"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_LEVEL_COLUMN).value
        #ダブルス
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_DOUBLES_COLUMN).value=="〇":
            df["ダブルス"]=True
        else:
            df["ダブルス"]=False
        #シングルス
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SINGLES_COLUMN).value=="〇":
            df["シングルス"]=True
        else:
            df["シングルス"]=False
        #ミックス
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_MIXED_COLUMN).value=="〇":
            df["ミックス"]=True
        else:
            df["ミックス"]=False
    else:
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_ATTENDANCE_COLUMN).value=="〇":
            df_cache["参加"]=True
        else:
            df_cache["参加"]=False
            
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SEX_COLUMN).value=="男":
            if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value.startswith("🔵")==False:
                df_cache["名前"]="🔵"+ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
            else:
                df_cache["名前"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
        else:
            if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value.startswith("🔴")==False:
                df_cache["名前"]="🔴"+ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
            else:
                df_cache["名前"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN).value
        
        df_cache["性別"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SEX_COLUMN).value
        df_cache["レベル"]=ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_LEVEL_COLUMN).value
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_DOUBLES_COLUMN).value=="〇":
            df_cache["ダブルス"]=True
        else:
            df_cache["ダブルス"]=False
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SINGLES_COLUMN).value=="〇":
            df_cache["シングルス"]=True
        else:
            df_cache["シングルス"]=False
        if ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_MIXED_COLUMN).value=="〇":
            df_cache["ミックス"]=True
        else:
            df_cache["ミックス"]=False
        df=pd.concat([df,df_cache])    
edited_df = st.data_editor(df,use_container_width=True,hide_index=True,height=None)

##########################################################################################################
#ボタン処理
##########################################################################################################

col1,col2=st.columns(2)
if col1.button("当日データの初期化",use_container_width=True):

    count=0
    #memberシートの初期化
    for row in ws_member:
        for cell in row :
            cell.value = None

    ws_member.cell(row=1,column=MEMBER_ATTENDANCE_COLUMN,value="参加")
    ws_member.cell(row=1,column=MEMBER_NAME_COLUMN,value="名前")
    ws_member.cell(row=1,column=MEMBER_SEX_COLUMN,value="性別")
    ws_member.cell(row=1,column=MEMBER_LEVEL_COLUMN,value="レベル")
    ws_member.cell(row=1,column=MEMBER_DOUBLES_COLUMN,value="ダブルス")
    ws_member.cell(row=1,column=MEMBER_SINGLES_COLUMN,value="シングルス")
    ws_member.cell(row=1,column=MEMBER_MIXED_COLUMN,value="ミックス")
    ws_member.cell(row=1,column=MEMBER_STATUS_COLUMN,value="ステータス")
    ws_member.cell(row=1,column=MEMBER_POINT_COLUMN,value="ポイント")
    ws_member.cell(row=1,column=MEMBER_NOT_COLUMN,value="回数")

    for number in range(len(df)):
        if edited_df.iat[number,0]==True:
            ws_member.cell(row=count+2,column=MEMBER_ATTENDANCE_COLUMN,value=edited_df.iat[number,0])#参加
            ws_member.cell(row=count+2,column=MEMBER_NAME_COLUMN,value=edited_df.iat[number,1])#名前
            ws_member.cell(row=count+2,column=MEMBER_SEX_COLUMN,value=edited_df.iat[number,2])#性別
            ws_member.cell(row=count+2,column=MEMBER_LEVEL_COLUMN,value=edited_df.iat[number,3])#レベル
            ws_member.cell(row=count+2,column=MEMBER_DOUBLES_COLUMN,value=edited_df.iat[number,4])#ダブルス
            ws_member.cell(row=count+2,column=MEMBER_SINGLES_COLUMN,value=edited_df.iat[number,5])#シングルス
            ws_member.cell(row=count+2,column=MEMBER_MIXED_COLUMN,value=edited_df.iat[number,6])#ミックス
            ws_member.cell(row=count+2,column=MEMBER_STATUS_COLUMN,value="待機")#ステータス
            ws_member.cell(row=count+2,column=MEMBER_POINT_COLUMN,value=0)#ポイント
            ws_member.cell(row=count+2,column=MEMBER_NOT_COLUMN,value=0)#回数
            count=count+1

    wb.save("data.xlsx")
    
    #コートの初期化
    for row in ws_court:
        for cell in row :
            cell.value = None

    ws_court.cell(row=1,column=1,value="勝者A")
    ws_court.cell(row=1,column=2,value="Aコート")
    ws_court.cell(row=1,column=3,value="勝者B")
    ws_court.cell(row=1,column=4,value="Bコート")
    ws_court.cell(row=1,column=5,value="勝者C")
    ws_court.cell(row=1,column=6,value="Cコート")
    wb.save("data.xlsx")
    
    #履歴の初期化
    for row in ws_history:
        for cell in row :
            cell.value = None

    ws_history.cell(row=1,column=1,value="名前1")
    ws_history.cell(row=1,column=2,value="名前2")
    ws_history.cell(row=1,column=3,value="名前3")
    ws_history.cell(row=1,column=4,value="名前4")
    ws_history.cell(row=1,column=5,value="コート")

    wb.save("data.xlsx")


if col2.button("データ更新",use_container_width=True):
    #base_parameterシートの初期化
    for row in ws_base_parameter:
        for cell in row :
            cell.value = None

    ws_base_parameter.cell(row=1,column=BASEPARAMETER_ATTENDANCE_COLUMN,value="参加")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_NAME_COLUMN,value="名前")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_SEX_COLUMN,value="性別")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_LEVEL_COLUMN,value="レベル")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_DOUBLES_COLUMN,value="ダブルス")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_SINGLES_COLUMN,value="シングルス")
    ws_base_parameter.cell(row=1,column=BASEPARAMETER_MIXED_COLUMN,value="ミックス")

    for number in range(len(df)):
        if edited_df.iat[number,0]==True:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_ATTENDANCE_COLUMN,value="〇")#参加
        else:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_ATTENDANCE_COLUMN,value="")#参加
        ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_NAME_COLUMN,value=edited_df.iat[number,1])#名前
        ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SEX_COLUMN,value=edited_df.iat[number,2])#性別
        ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_LEVEL_COLUMN,value=edited_df.iat[number,3])#レベル
        if edited_df.iat[number,4]==True:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_DOUBLES_COLUMN,value="〇")#ダブルス
        else:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_DOUBLES_COLUMN,value="")#ダブルス
        if edited_df.iat[number,5]==True:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SINGLES_COLUMN,value="〇")#シングルス
        else:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_SINGLES_COLUMN,value="")#シングルス
        if edited_df.iat[number,6]==True:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_MIXED_COLUMN,value="〇")#ミックス
        else:
            ws_base_parameter.cell(row=number+2,column=BASEPARAMETER_MIXED_COLUMN,value="")#ミックス

    wb.save("data.xlsx")


##########################################################################################################
#説明
##########################################################################################################
"[使い方]"
"当日初期に来る可能性があるメンバの参加状況を確認し、「データ更新ボタン」と「参加者を更新ボタン」を押す。"
"以降は基本いじらない事"

"[説明]"
"管理用のエクセル(base_parameter)が表示されている。操作しただけでは何も変わらない。(画面が変わるだけ)"
"「参加者の更新」ボタンを押すと、当日のメンバーリスト(member)とコート情報(court)と履歴(history)が初期化される。(ステータスやポイントもリセットされる)"
"「データ更新」ボタンを押すと現在表示されている内容を管理用のエクセル(base_parameter)に書き戻す。"
"データ更新せずに画面遷移した場合、戻ってくると初期状態に戻っている。"


