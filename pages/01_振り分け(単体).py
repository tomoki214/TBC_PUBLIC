
""
"[使い方]"
"待機は4名～6名程度。(休憩最大2名の想定)"

#
#######################################################################
#一度コートに入れるとコートから出せない。
#メンバの更新ができない。
#######################################################################
import streamlit as st
from streamlit_sortables import sort_items
import pandas as pd

#エクセル操作
import openpyxl

wb=openpyxl.load_workbook("data.xlsx",data_only=True)
ws_base_parameter=wb["base_parameter"]
ws_member=wb["member"]
ws_history=wb["history"]
ws_court=wb["court"]

##########################################################################################################
#定義
##########################################################################################################

stanby_member_original=[
    {'header':'待機','items':[]},
    {'header':'休憩','items':[]},
]
custom_style = """
.sortable-component {
    border: 3px solid #7d7d7d;
    font-color: #FFFFFF;
    border-radius: 10px;
    padding: 0px;
}
.sortable-container {
    background-color: #7d7d7d;
    font-color: #FFFFFF;
    counter-reset: item;
}
.sortable-container-header {
    background-color: #CCCCCC;
    font-color: #FFFFFF;
    padding-left: 1rem;
}
.sortable-container-body {
    background-color: #FFFFFF;
    font-color: #FFFFFF;
}
.sortable-item, .sortable-item:hover {
    background-color: #7d7d7d;
    font-color: #FFFFFF;
    font-weight: bold;
}
"""
###########################################
#memberのシートを回数順にソートする
###########################################
def member_sort():
    for i in range(30):
        if ws_member.cell(row=i+1,column=2).value==None:
            break
        else:
            min_value=ws_member.cell(row=i+2,column=8).value

        for number in range(30):
            j=number+i
            if ws_member.cell(row=j+2,column=2).value==None:
                break
            else:
                if ws_member.cell(row=j+2,column=8).value<min_value:
                    swap1=ws_member.cell(row=j+2,column=1).value
                    swap2=ws_member.cell(row=j+2,column=2).value
                    swap3=ws_member.cell(row=j+2,column=3).value
                    swap4=ws_member.cell(row=j+2,column=4).value
                    swap5=ws_member.cell(row=j+2,column=5).value
                    swap6=ws_member.cell(row=j+2,column=6).value
                    swap7=ws_member.cell(row=j+2,column=7).value
                    swap8=ws_member.cell(row=j+2,column=8).value
                    swap9=ws_member.cell(row=j+2,column=9).value
                    ws_member.cell(row=j+2,column=1,value=ws_member.cell(row=i+2,column=1).value)
                    ws_member.cell(row=j+2,column=2,value=ws_member.cell(row=i+2,column=2).value)
                    ws_member.cell(row=j+2,column=3,value=ws_member.cell(row=i+2,column=3).value)
                    ws_member.cell(row=j+2,column=4,value=ws_member.cell(row=i+2,column=4).value)
                    ws_member.cell(row=j+2,column=5,value=ws_member.cell(row=i+2,column=5).value)
                    ws_member.cell(row=j+2,column=6,value=ws_member.cell(row=i+2,column=6).value)
                    ws_member.cell(row=j+2,column=7,value=ws_member.cell(row=i+2,column=7).value)
                    ws_member.cell(row=j+2,column=8,value=ws_member.cell(row=i+2,column=8).value)
                    ws_member.cell(row=j+2,column=9,value=ws_member.cell(row=i+2,column=9).value)
                    ws_member.cell(row=i+2,column=1,value=swap1)
                    ws_member.cell(row=i+2,column=2,value=swap2)
                    ws_member.cell(row=i+2,column=3,value=swap3)
                    ws_member.cell(row=i+2,column=4,value=swap4)
                    ws_member.cell(row=i+2,column=5,value=swap5)
                    ws_member.cell(row=i+2,column=6,value=swap6)
                    ws_member.cell(row=i+2,column=7,value=swap7)
                    ws_member.cell(row=i+2,column=8,value=swap8)
                    ws_member.cell(row=i+2,column=9,value=swap9)
                    min_value=swap8
    wb.save("data.xlsx")

###########################################
#ステータスの最新化
###########################################
def status_reset():
    for number in range(30):
        if ws_member.cell(row=(number+1),column=2).value==None:
            break
        #一旦待機にして、間違っていれば上書きする
        ws_member.cell(row=(number+1),column=7,value="待機")

        #Aコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=1).value==ws_member.cell(row=(number+1),column=2).value:
                ws_member.cell(row=(number+1),column=7,value="Aコート")
        #Bコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=2).value==ws_member.cell(row=(number+1),column=2).value:
                ws_member.cell(row=(number+1),column=7,value="Bコート")
        #Cコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=3).value==ws_member.cell(row=(number+1),column=2).value:
                ws_member.cell(row=(number+1),column=7,value="Cコート")

        #休憩か確認
        for number2 in range(len(st.session_state["sorted_data"][1]["items"])):
            if st.session_state["sorted_data"][1]["items"][number2]==ws_member.cell(row=(number+1),column=2).value:
                ws_member.cell(row=(number+1),column=7,value="休憩")
    wb.save("data.xlsx")

###########################################
#nameの人のステータスをstatusに変更する
###########################################
def status_update(name,status):
    for number in range(30):
        if ws_member.cell(row=(number+1),column=2).value==name:
            ws_member.cell(row=(number+1),column=7,value=status)
            break
    wb.save("data.xlsx")

###########################################
#nameの人にポイントにpointを加算する
###########################################
def point_add(name,point):
    for number in range(30):
        if ws_member.cell(row=(number+1),column=2).value==name:
            if ws_member.cell(row=(number+1),column=8).value==None:
               ws_member.cell(row=(number+1),column=8,value=point)
            else:
                ws_member.cell(row=(number+1),column=8,value=point+ws_member.cell(row=(number+1),column=8).value)

            if ws_member.cell(row=(number+1),column=9).value==None:
                ws_member.cell(row=(number+1),column=9,value=1)
            else:
                ws_member.cell(row=(number+1),column=9,value=1+ws_member.cell(row=(number+1),column=9).value)
            break
    wb.save("data.xlsx")

###########################################
#履歴に名前とコートを記載する
###########################################
def history_add(name1,name2,name3,name4,court):
    for number in range(200):
        if ws_history.cell(row=(number+1),column=2).value==None:
            ws_history.cell(row=(number+1),column=1,value=name1)
            ws_history.cell(row=(number+1),column=2,value=name2)
            ws_history.cell(row=(number+1),column=3,value=name3)
            ws_history.cell(row=(number+1),column=4,value=name4)
            ws_history.cell(row=(number+1),column=5,value=court)
            break
    wb.save("data.xlsx")

###########################################
#名前からレベルを調べる
###########################################
def search_level(name):
    if name==None:
        return 0
    for number in range(30):
        if ws_member.cell(row=(number+2),column=2).value==name:
            return ws_member.cell(row=(number+2),column=4).value   #レベルを返す

    return 0

##########################################################################################################
#コート
##########################################################################################################

import copy

def court_sorting(court_name,court_number):
   count=0
   ave_level=0
   if ws_court.cell(row=2,column=court_number).value!="":
        #コートのメンバを待機にして平均レベルを算出する。
        name_list=[]
        for number in range(4):
            name_list=name_list+[ws_court.cell(row=number+2,column=court_number).value]    #コートのメンバの名前を抽出
            status_update(name_list[number],"待機")                                         #ステータスを変更する
            ave_level=ave_level+search_level(name_list[number])                             #レベルを出す
            if name_list[number]!=None:
                count=count+1
        
        #結果的にシングルスの場合は2で割らない=2倍なので処理しない
        if count==4:
            ave_level=ave_level/count
        
        #ポイント追加
        for number in range(4):
            if search_level(name_list[number])!=0:
                point_add(name_list[number],ave_level/search_level(name_list[number]))          #ポイントを追加する
        #履歴追加
        history_add(name_list[0],name_list[1],name_list[2],name_list[3],court_name)         #履歴に記録する
        #シングルスに対応していない

   
   member_sort()    #ポイントが少ない人を上にする

   #変数定義
   name_list_candidate=["","","",""]   #候補者の名前保存
   point_min=1000              #候補者のポイント保存


    ###################################
    ###レベル4男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="男":
                if ws_member.cell(row=number+2,column=4).value>=3.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###レベル4女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="女":
                if ws_member.cell(row=number+2,column=4).value>=3.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###レベル3男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="男":
                if ws_member.cell(row=number+2,column=4).value>=2.5:
                    if ws_member.cell(row=number+2,column=4).value<=3.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###レベル3女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="女":
                if ws_member.cell(row=number+2,column=4).value>=2.5:
                    if ws_member.cell(row=number+2,column=4).value<=3.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###レベル2男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="男":
                if ws_member.cell(row=number+2,column=4).value<=2.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###レベル2女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if point_cache<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=3).value=="女":
                if ws_member.cell(row=number+2,column=4).value<=2.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###レベル4ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if point_cache<point_min:
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="男":
                    if ws_member.cell(row=number+2,column=4).value>=3.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count1=count1+1
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="女":
                    if ws_member.cell(row=number+2,column=4).value>=3.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count2=count2+1

    ###################################
    ###レベル3ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if point_cache<point_min:
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="男":
                    if ws_member.cell(row=number+2,column=4).value>=2.5:
                        if ws_member.cell(row=number+2,column=4).value<=3.5:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                            count1=count1+1
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="女":
                    if ws_member.cell(row=number+2,column=4).value>=2.5:
                        if ws_member.cell(row=number+2,column=4).value<=3.5:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                            count2=count2+1

    ###################################
    ###レベル2ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if point_cache<point_min:
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="男":
                    if ws_member.cell(row=number+2,column=4).value<=2.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count1=count1+1
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=6).value==1:
                if ws_member.cell(row=number+2,column=3).value=="女":
                    if ws_member.cell(row=number+2,column=4).value<=2.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count2=count2+1

    ###################################
    ###レベル4シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if (point_cache*2)<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=5).value==1:
                if ws_member.cell(row=number+2,column=4).value>=3.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###レベル3シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if (point_cache*2)<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=5).value==1:
                if ws_member.cell(row=number+2,column=4).value<=3.5:
                    if ws_member.cell(row=number+2,column=4).value>=2.5:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###レベル2シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if (point_cache*2)<point_min:
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=5).value==1:
                if ws_member.cell(row=number+2,column=4).value<=2.5:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###指導バドの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント
   count=0
   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if point_cache<point_min:
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=4).value>=2.5:
                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                count1=count1+1
        if ws_member.cell(row=number+2,column=7).value=="待機":
            if ws_member.cell(row=number+2,column=4).value<=1.5:
                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                count2=count2+1
    

    ###################################
    ###全体のポイント差を計算する
    ###################################
   all_point_max=0
   all_point_min=1000

   for number in range(30):
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            all_point_min=ws_member.cell(row=number+2,column=8).value
            break

   for number in range(30):
        if ws_member.cell(row=number+2,column=4).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value=="待機":
            all_point_max=ws_member.cell(row=number+2,column=8).value

    ###################################
    ###回数のみ
    ###################################

   if point_min==1000 or all_point_max-all_point_min>=2:
        count=0                  #4名カウントするための数字
        name_list_cache=[]       #一時的な名前保存
        point_cache=0            #一時的なポイント保存
        for number in range(30):
                #キャッシュデータの更新
                if count>=4:
                    if point_cache<point_min:
                        point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                        name_list_candidate=copy.deepcopy(name_list_cache)
                        break
                if ws_member.cell(row=number+2,column=4).value==None:
                    break
                if ws_member.cell(row=number+2,column=7).value=="待機":
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=2).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=8).value                         #ポイントを加算
                    count=count+1


    #決定した条件でシートを更新する。
   for number in range(4):
        if len(name_list_candidate)>number:
            status_update(name_list_candidate[number],court_name)                                       #memberシートの更新                      
            ws_court.cell(row=number+2,column=court_number,value=name_list_candidate[number])            #courtシートの更新
        else:
            #シングルスの場合残りはブランクにする
            ws_court.cell(row=number+2,column=court_number,value="")            #courtシートの更新
   wb.save("data.xlsx")


def court_clear(court_name,court_number):
   
    #対象コートを空にする。
    for number in range(4):
        ws_court.cell(row=number+2,column=court_number,value="")

    #対象コートに入っていたメンバのステータスを待機にする
    for number in range(30):
        if ws_member.cell(row=number+2,column=7).value==None:
            break
        if ws_member.cell(row=number+2,column=7).value==court_name:
            ws_member.cell(row=number+2,column=7,value="待機")
    wb.save("data.xlsx")
                           

col1,col2,col3=st.columns(3)
if col1.button("Aコート",use_container_width=True):
    court_sorting("Aコート",1)
if col2.button("Bコート",use_container_width=True):
    court_sorting("Bコート",2)
if col3.button("Cコート",use_container_width=True):
    court_sorting("Cコート",3)

col4,col5,col6=st.columns(3)
if col4.button("Aコートを空にする",use_container_width=True):
    court_clear("Aコート",1)
if col5.button("Bコートを空にする",use_container_width=True):
    court_clear("Bコート",2)
if col6.button("Cコートを空にする",use_container_width=True):
    court_clear("Cコート",3)

df = pd.read_excel(r'data.xlsx',sheet_name='court')
st.table(df)



##########################################################################################################
#StandByとBreakのインターフェイス
##########################################################################################################
stanby_member_original[0]["items"].clear()
stanby_member_original[1]["items"].clear()

for number in range(30):
    if ws_member.cell(row=(number+2),column=2).value!=None:
        if(ws_member.cell(row=(number+2),column=7).value=="待機"):
            stanby_member_original[0]["items"].append(ws_member.cell(row=(number+2),column=2).value)
        if(ws_member.cell(row=(number+2),column=7).value=="休憩"):
            stanby_member_original[1]["items"].append(ws_member.cell(row=(number+2),column=2).value)
    else:
        break

st.session_state["sorted_data"]=sort_items(stanby_member_original, multi_containers=True,custom_style=custom_style)

#休憩にデータが入ったらエクセルも更新する
for count in range(len(st.session_state["sorted_data"][1]["items"])):
    for number in range(30):
        if ws_member.cell(row=(number+2),column=2).value!=None:
            if(ws_member.cell(row=(number+2),column=2).value==st.session_state["sorted_data"][1]["items"][count]):
                ws_member.cell(row=(number+2),column=7,value="休憩")
                break

#待機にデータが入ったらエクセルも更新する
for count in range(len(st.session_state["sorted_data"][0]["items"])):
    for number in range(30):
        if ws_member.cell(row=(number+2),column=2).value!=None:
            if(ws_member.cell(row=(number+2),column=2).value==st.session_state["sorted_data"][0]["items"][count]):
                ws_member.cell(row=(number+2),column=7,value="待機")
                break

wb.save("data.xlsx")



##########################################################################################################
#履歴
##########################################################################################################

if st.button("履歴の削除"):
    for row in ws_history:
        for cell in row :
            cell.value = None
    ws_history.cell(row=1,column=1,value="名前1")
    ws_history.cell(row=1,column=2,value="名前2")
    ws_history.cell(row=1,column=3,value="名前3")
    ws_history.cell(row=1,column=4,value="名前4")
    ws_history.cell(row=1,column=5,value="コート")

    wb.save("data.xlsx")

df = pd.read_excel(r'data.xlsx',sheet_name='history')

st.table(df)