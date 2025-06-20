
""
"※使い方が分からない、エラーが出た場合は操作せず、一旦フリーコートとしてコートを活用しておいてください。"
"[使い方]"
"・ゲームが終了したら勝ったほうにチェックを入れ、終わったコートのボタンを押すこと"
"・無料サーバを使っている関係で処理が遅くなるが連打しないこと。"

#
#######################################################################
#一度コートに入れるとコートから出せない。
#メンバの更新ができない。
#######################################################################
import streamlit as st
from streamlit_sortables import sort_items
import pandas as pd

#エクセル操作
import openpyxl

wb=openpyxl.load_workbook("data.xlsx",data_only=True)
ws_base_parameter=wb["base_parameter"]
ws_member=wb["member"]
ws_history=wb["history"]
ws_court=wb["court"]

##########################################################################################################
#定義
##########################################################################################################

#############################
#ベースパラメータシートの列定義
#############################
BASEPARAMETER_ATTENDANCE_COLUMN=1
BASEPARAMETER_NAME_COLUMN=2
BASEPARAMETER_SEX_COLUMN=3
BASEPARAMETER_LEVEL_COLUMN=4
BASEPARAMETER_DOUBLES_COLUMN=5
BASEPARAMETER_SINGLES_COLUMN=6
BASEPARAMETER_MIXED_COLUMN=7

#############################
#メンバーシートの列定義
#############################
MEMBER_ATTENDANCE_COLUMN=1      #置換完了
MEMBER_NAME_COLUMN=2            #置換完了
MEMBER_SEX_COLUMN=3             #途中
MEMBER_LEVEL_COLUMN=4
MEMBER_DOUBLES_COLUMN=5
MEMBER_SINGLES_COLUMN=6
MEMBER_MIXED_COLUMN=7
MEMBER_STATUS_COLUMN=8
MEMBER_POINT_COLUMN=9
MEMBER_NOT_COLUMN=10

#############################
#ヒストリーシートの列定義
#############################
HISTORY_NAME1_COLUMN=1
HISTORY_NAME2_COLUMN=2
HISTORY_NAME3_COLUMN=3
HISTORY_NAME4_COLUMN=4
HISTORY_COURT_COLUMN=5

#############################
#コートシートの列定義(変数で使われているところが多いため、ここの編集だけでは不可)
#############################
COURT_WINLOSEA_COLUMN=1
COURT_NAMEA_COLUMN=2
COURT_WINLOSEB_COLUMN=3
COURT_NAMEB_COLUMN=4
COURT_WINLOSEC_COLUMN=5
COURT_NAMEC_COLUMN=6


ADJUSTMENT_NUMBER=10        #優先された場合に利用
LOSE_ADD_POINT=0.3          #負けた時の加算ポイント
CAL_ADD_POINT=1             #中間レベルの開催ロジックの傾斜

ALL_CAL_ADD_POINT1=3        #回数差が出過ぎた時の調整用ポイント1
ALL_CAL_ADD_POINT2=6        #回数差が出過ぎた時の調整用ポイント2

stanby_member_original=[
    {'header':'待機','items':[]},
    {'header':'休憩','items':[]},
]
custom_style = """
.sortable-component {
    border: 3px solid #7d7d7d;
    font-color: #FFFFFF;
    border-radius: 10px;
    padding: 0px;
}
.sortable-container {
    background-color: #7d7d7d;
    font-color: #FFFFFF;
    counter-reset: item;
}
.sortable-container-header {
    background-color: #CCCCCC;
    font-color: #FFFFFF;
    padding-left: 1rem;
}
.sortable-container-body {
    background-color: #FFFFFF;
    font-color: #FFFFFF;
}
.sortable-item, .sortable-item:hover {
    background-color: #7d7d7d;
    font-color: #FFFFFF;
    font-weight: bold;
}
"""
###########################################
#memberのシートを回数順にソートする
###########################################
def member_sort():
    for i in range(30):
        if ws_member.cell(row=i+1,column=MEMBER_NAME_COLUMN).value==None:
            break
        else:
            min_value=ws_member.cell(row=i+2,column=MEMBER_POINT_COLUMN).value    #最小のポイントを返す

        for number in range(30):
            j=number+i
            if ws_member.cell(row=j+2,column=MEMBER_NAME_COLUMN).value==None:
                break
            else:
                if ws_member.cell(row=j+2,column=MEMBER_POINT_COLUMN).value<min_value:
                    swap1=ws_member.cell(row=j+2,column=MEMBER_ATTENDANCE_COLUMN).value
                    swap2=ws_member.cell(row=j+2,column=MEMBER_NAME_COLUMN).value
                    swap3=ws_member.cell(row=j+2,column=MEMBER_SEX_COLUMN).value
                    swap4=ws_member.cell(row=j+2,column=MEMBER_LEVEL_COLUMN).value
                    swap5=ws_member.cell(row=j+2,column=MEMBER_DOUBLES_COLUMN).value
                    swap6=ws_member.cell(row=j+2,column=MEMBER_SINGLES_COLUMN).value
                    swap7=ws_member.cell(row=j+2,column=MEMBER_MIXED_COLUMN).value
                    swap8=ws_member.cell(row=j+2,column=MEMBER_STATUS_COLUMN).value
                    swap9=ws_member.cell(row=j+2,column=MEMBER_POINT_COLUMN).value
                    swap10=ws_member.cell(row=j+2,column=MEMBER_NOT_COLUMN).value
                    
                    ws_member.cell(row=j+2,column=MEMBER_ATTENDANCE_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_ATTENDANCE_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_NAME_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_NAME_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_SEX_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_SEX_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_LEVEL_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_LEVEL_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_DOUBLES_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_DOUBLES_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_SINGLES_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_SINGLES_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_MIXED_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_MIXED_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_STATUS_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_STATUS_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_POINT_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_POINT_COLUMN).value)
                    ws_member.cell(row=j+2,column=MEMBER_NOT_COLUMN,value=ws_member.cell(row=i+2,column=MEMBER_NOT_COLUMN).value)

                    ws_member.cell(row=i+2,column=MEMBER_ATTENDANCE_COLUMN,value=swap1)
                    ws_member.cell(row=i+2,column=MEMBER_NAME_COLUMN,value=swap2)
                    ws_member.cell(row=i+2,column=MEMBER_SEX_COLUMN,value=swap3)
                    ws_member.cell(row=i+2,column=MEMBER_LEVEL_COLUMN,value=swap4)
                    ws_member.cell(row=i+2,column=MEMBER_DOUBLES_COLUMN,value=swap5)
                    ws_member.cell(row=i+2,column=MEMBER_SINGLES_COLUMN,value=swap6)
                    ws_member.cell(row=i+2,column=MEMBER_MIXED_COLUMN,value=swap7)
                    ws_member.cell(row=i+2,column=MEMBER_STATUS_COLUMN,value=swap8)
                    ws_member.cell(row=i+2,column=MEMBER_POINT_COLUMN,value=swap9)
                    ws_member.cell(row=i+2,column=MEMBER_NOT_COLUMN,value=swap10)
                    min_value=swap9
    wb.save("data.xlsx")

###########################################
#ステータスの最新化
###########################################
def status_reset():
    for number in range(30):
        if ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value==None:
            break
        #一旦待機にして、間違っていれば上書きする
        ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value="待機")

        #Aコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=COURT_NAMEA_COLUMN).value==ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value:
                ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value="Aコート")
        #Bコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=COURT_NAMEB_COLUMN).value==ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value:
                ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value="Bコート")
        #Cコートか確認
        for number2 in range(4):
            if ws_court.cell(row=(number2+1),column=COURT_NAMEC_COLUMN).value==ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value:
                ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value="Cコート")

        #休憩か確認
        for number2 in range(len(st.session_state["sorted_data"][1]["items"])):
            if st.session_state["sorted_data"][1]["items"][number2]==ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value:
                ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value="休憩")
    wb.save("data.xlsx")

###########################################
#nameの人のステータスをstatusに変更する
###########################################
def status_update(name,status):
    for number in range(30):
        if ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value==name:
            ws_member.cell(row=(number+1),column=MEMBER_STATUS_COLUMN,value=status)
            break
    wb.save("data.xlsx")

###########################################
#nameの人にポイントにpointを加算する
###########################################
def point_add(name,point):
    for number in range(30):
        if ws_member.cell(row=(number+1),column=MEMBER_NAME_COLUMN).value==name:
            if ws_member.cell(row=(number+1),column=MEMBER_POINT_COLUMN).value==None:
               ws_member.cell(row=(number+1),column=MEMBER_POINT_COLUMN,value=point)
            else:
                ws_member.cell(row=(number+1),column=MEMBER_POINT_COLUMN,value=point+ws_member.cell(row=(number+1),column=MEMBER_POINT_COLUMN).value)

            if ws_member.cell(row=(number+1),column=MEMBER_NOT_COLUMN).value==None:
                ws_member.cell(row=(number+1),column=MEMBER_NOT_COLUMN,value=1)
            else:
                ws_member.cell(row=(number+1),column=MEMBER_NOT_COLUMN,value=1+ws_member.cell(row=(number+1),column=MEMBER_NOT_COLUMN).value)
            break
    wb.save("data.xlsx")

###########################################
#履歴に名前とコートを記載する
###########################################
def history_add(name1,name2,name3,name4,court):
    for number in range(200):
        if ws_history.cell(row=(number+1),column=HISTORY_NAME1_COLUMN).value==None:
            ws_history.cell(row=(number+1),column=HISTORY_NAME1_COLUMN,value=name1)
            ws_history.cell(row=(number+1),column=HISTORY_NAME2_COLUMN,value=name2)
            ws_history.cell(row=(number+1),column=HISTORY_NAME3_COLUMN,value=name3)
            ws_history.cell(row=(number+1),column=HISTORY_NAME4_COLUMN,value=name4)
            ws_history.cell(row=(number+1),column=HISTORY_COURT_COLUMN,value=court)
            break
    wb.save("data.xlsx")

###########################################
#名前からレベルを調べる
###########################################
def search_level(name):
    if name==None:
        return 0
    for number in range(30):
        if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==name:
            return ws_member.cell(row=(number+2),column=MEMBER_LEVEL_COLUMN).value   #レベルを返す

    return 0

###########################################
#名前から勝敗を調べる
#勝っていればTrue、負けていればFalse、名前がなければNoneの想定。
###########################################
def search_win(name):
    if name==None:
        return 0
    for number in range(4):
        for number2 in range(3):
            if ws_court.cell(row=(number+2),column=(number2+1)*2).value==name:
                return ws_court.cell(row=(number+2),column=(number2+1)*2-1).value #勝敗のパラメータを返す。
            
    return None

##########################################################################################################
#コート
##########################################################################################################

import copy

def court_sorting(court_name,court_number):
   count=0
   ave_level=0
   court_number=court_number*2

   if ws_court.cell(row=2,column=court_number).value!="":
        #コートのメンバを待機にして平均レベルを算出する。
        name_list=[]
        for number in range(4):
            name_list=name_list+[ws_court.cell(row=number+2,column=court_number).value]    #コートのメンバの名前を抽出
            status_update(name_list[number],"待機")                                         #ステータスを変更する
            ave_level=ave_level+search_level(name_list[number])                             #レベルを出す
            if name_list[number]!=None:
                count=count+1
        
        #結果的にシングルスの場合は2で割らない=2倍なので処理しない
        if count==4:
            ave_level=ave_level/4
        else:
            ave_level=ave_level/2
        
        #ポイント追加
        for number in range(4):
            if search_level(name_list[number])!=0:
                if search_win(name_list[number])==True:
                    point_add(name_list[number],ave_level/search_level(name_list[number]))    #勝っていたらポイントが少なく計上される。
                else:
                    point_add(name_list[number],ave_level/search_level(name_list[number])+LOSE_ADD_POINT)          #ポイントを追加する
        #履歴追加
        history_add(name_list[0],name_list[1],name_list[2],name_list[3],court_name)         #履歴に記録する

   member_sort()    #ポイントが少ない人を上にする

   #変数定義
   name_list_candidate=["","","",""]   #候補者の名前保存
   point_min=1000              #候補者のポイント保存

   court_level=""

    ###################################
    ###上級男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="上級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.8:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###上級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="上級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.8:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###中級男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="中級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1

    ###################################
    ###中級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="中級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1

    ###################################
    ###初級男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1

    ###################################
    ###初級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1

    ###################################
    ###初心者ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初心者/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=2.7:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###上級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="上級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.8:
                        if count1<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.8:
                        if count2<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count2=count2+1

    ###################################
    ###中級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="中級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            if count1<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            if count2<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count2=count2+1

    ###################################
    ###初級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="初級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                            if count1<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                            if count2<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count2=count2+1

    ###################################
    ###中/上級男子ダブルスの選定
    ###################################
    
   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="中上級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###中/上級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="中上級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1


    ###################################
    ###初/中級男子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初中級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1

    ###################################
    ###初/中級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初中級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count=count+1


    ###################################
    ###初/初級男子ダブルスの選定
    ###################################
    
   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初初級/男子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###初/初級女子ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="ダブルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if point_cache<point_min:
                court_level="初初級/女子/ダブルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1


    ###################################
    ###中/上級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                point_cache=point_cache+CAL_ADD_POINT
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="中上級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if count1<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        if count2<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count2=count2+1

    ###################################
    ###初/中級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                point_cache=point_cache+CAL_ADD_POINT
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="初中級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            if count1<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                            if count2<2:
                                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                                count2=count2+1

    ###################################
    ###初/初級ミックスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                point_cache=point_cache+CAL_ADD_POINT
                if mode=="ミックス":
                    point_cache=point_cache-ADJUSTMENT_NUMBER
                if point_cache<point_min:
                    court_level="初初級/ミックス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="男":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                        if count1<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_MIXED_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_SEX_COLUMN).value=="女":
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                        if count2<2:
                            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                            count2=count2+1



    ###################################
    ###上級シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if mode=="シングルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if (point_cache*2)<point_min:
                court_level="上級/シングルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.8:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###中級シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if mode=="シングルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if (point_cache*2)<point_min:
                court_level="中級/シングルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1

    ###################################
    ###初級シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            if mode=="シングルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if (point_cache*2)<point_min:
                court_level="初級/シングルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.2:
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1


    ###################################
    ###中/上級シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="シングルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if (point_cache*2)<point_min:
                court_level="中上級/シングルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                    count=count+1

    ###################################
    ###初/中級シングルスの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count=0                 #男のカウント

   for number in range(30):
        #キャッシュデータの更新
        if count>=2:
            point_cache=point_cache+CAL_ADD_POINT
            if mode=="シングルス":
                point_cache=point_cache-ADJUSTMENT_NUMBER
            if (point_cache*2)<point_min:
                court_level="初中級/シングルス"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_SINGLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=2.8:
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.7:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1



    ###################################
    ###全体(上)ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+ALL_CAL_ADD_POINT1
            if point_cache<point_min:
                court_level="全体/ダブルス1"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=2.8:
                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                count=count+1

    ###################################
    ###全体(下)ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+ALL_CAL_ADD_POINT1
            if point_cache<point_min:
                court_level="全体/ダブルス2"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                count=count+1

    ###################################
    ###全体ダブルスの選定
    ###################################

   count=0                  #4名カウントするための数字
   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   for number in range(30):
        #キャッシュデータの更新
        if count>=4:
            point_cache=point_cache+ALL_CAL_ADD_POINT2
            if point_cache<point_min:
                court_level="全体/ダブルス3"
                point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                name_list_candidate=copy.deepcopy(name_list_cache)
                break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
            point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
            count=count+1

    ###################################
    ###指導バドの選定
    ###################################

   name_list_cache=[]       #一時的な名前保存
   point_cache=0            #一時的なポイント保存
   count1=0                 #男のカウント
   count2=0                 #女のカウント
   count=0
   for number in range(30):
        #キャッシュデータの更新
        if count1>=2:
            if count2>=2:
                point_cache=point_cache+CAL_ADD_POINT
                if point_cache<point_min:
                    court_level="指導/ダブルス"
                    point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                    name_list_candidate=copy.deepcopy(name_list_cache)
                    break
        if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value>=3.3:
                    if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=3.7:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count1=count1+1
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
            if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value<=2.7:
                    name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                    point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                    count2=count2+1
    
    ###################################
    ###回数のみ
    ###################################

   if mode=="回数優先":
        count=0                  #4名カウントするための数字
        name_list_cache=[]       #一時的な名前保存
        point_cache=0            #一時的なポイント保存
        for number in range(30):
                #キャッシュデータの更新
                if count>=4:
                    if point_cache<point_min:
                        court_level="回数/ダブルス"
                        point_min=point_cache                                                                       #規定人数集まった場合だけ更新するか判断
                        name_list_candidate=copy.deepcopy(name_list_cache)
                        break
                if ws_member.cell(row=number+2,column=MEMBER_LEVEL_COLUMN).value==None:
                    break
                if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value=="待機":
                    if ws_member.cell(row=number+2,column=MEMBER_DOUBLES_COLUMN).value==1:
                        name_list_cache=name_list_cache+[ws_member.cell(row=number+2,column=MEMBER_NAME_COLUMN).value]              #名前をリストに追加     
                        point_cache=point_cache+ws_member.cell(row=number+2,column=MEMBER_POINT_COLUMN).value                         #ポイントを加算
                        count=count+1


    #決定した条件でシートを更新する。
   for number in range(4):
        if len(name_list_candidate)>number:
            status_update(name_list_candidate[number],court_name)                                       #memberシートの更新                      
            ws_court.cell(row=number+2,column=court_number,value=name_list_candidate[number])            #courtシートの更新
            ws_court.cell(row=number+2,column=court_number-1,value=False)                               #★ここで左隣をfalseにしたい。
        else:
            #シングルスの場合残りはブランクにする
            ws_court.cell(row=number+2,column=court_number,value="")            #courtシートの更新
            ws_court.cell(row=number+2,column=court_number-1,value=False)       #★ここで左隣をfalseにしたい。
   wb.save("data.xlsx")
   st.write(court_level)


def court_clear(court_name,court_number):
    court_number=court_number*2                                                                         #★列をずらす
    #対象コートを空にする。
    for number in range(4):
        ws_court.cell(row=number+2,column=court_number,value="")
        ws_court.cell(row=number+2,column=court_number-1,value=False)                                   #★勝敗も削除

    #対象コートに入っていたメンバのステータスを待機にする
    for number in range(30):
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value==None:
            break
        if ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN).value==court_name:
            ws_member.cell(row=number+2,column=MEMBER_STATUS_COLUMN,value="待機")
    wb.save("data.xlsx")

##########################################################################################################
#コートのインターフェイス
##########################################################################################################

df = pd.DataFrame([{}])
df_cache = pd.DataFrame([{}])

mode=st.selectbox('アクションを選択してください。',['ランダム','ダブルス','ミックス','シングルス','回数優先','[コートを空にする]'])

#変数を定義/初期化
if 'winloseA' not in st.session_state:
    st.session_state.winloseA=[False,False,False,False]
    st.session_state.winloseB=[False,False,False,False]
    st.session_state.winloseC=[False,False,False,False]

col1,col2,col3=st.columns(3)
if col1.button("Aコート",use_container_width=True):
    #勝敗をエクセルに書き込む
    for number in range(4):
        ws_court.cell(row=number+2,column=COURT_WINLOSEA_COLUMN,value=st.session_state.winloseA[number])        #Aコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEB_COLUMN,value=st.session_state.winloseB[number])        #Bコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEC_COLUMN,value=st.session_state.winloseC[number])        #Cコート

    #待機/休憩の状態を変数からエクセルに書き込む
    #メンバの回数だけmemerシートを見てステータスを更新する。
    #待機
    for number2 in range(st.session_state.stanby_member0_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member0[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="待機")
                    break
    #待機
    for number2 in range(st.session_state.stanby_member1_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member1[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="休憩")
                    break
    wb.save("data.xlsx")

    if mode=="[コートを空にする]":
        court_clear("Aコート",1)
    else:
        court_sorting("Aコート",1)
if col2.button("Bコート",use_container_width=True):
    #勝敗をエクセルに書き込む
    for number in range(4):
        ws_court.cell(row=number+2,column=COURT_WINLOSEA_COLUMN,value=st.session_state.winloseA[number])        #Aコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEB_COLUMN,value=st.session_state.winloseB[number])        #Bコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEC_COLUMN,value=st.session_state.winloseC[number])        #Cコート
    #待機/休憩の状態を変数からエクセルに書き込む
    #メンバの回数だけmemerシートを見てステータスを更新する。
    #待機
    for number2 in range(st.session_state.stanby_member0_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member0[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="待機")
                    break
    #待機
    for number2 in range(st.session_state.stanby_member1_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member1[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="休憩")
                    break
    wb.save("data.xlsx")

    if mode=="[コートを空にする]":
        court_clear("Bコート",2)
    else:
        court_sorting("Bコート",2)
if col3.button("Cコート",use_container_width=True):
    #勝敗をエクセルに書き込む
    for number in range(4):
        ws_court.cell(row=number+2,column=COURT_WINLOSEA_COLUMN,value=st.session_state.winloseA[number])        #Aコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEB_COLUMN,value=st.session_state.winloseB[number])        #Bコート
        ws_court.cell(row=number+2,column=COURT_WINLOSEC_COLUMN,value=st.session_state.winloseC[number])        #Cコート
    #待機/休憩の状態を変数からエクセルに書き込む
    #メンバの回数だけmemerシートを見てステータスを更新する。
    #待機
    for number2 in range(st.session_state.stanby_member0_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member0[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="待機")
                    break
    #待機
    for number2 in range(st.session_state.stanby_member1_num):
        for number in range(30):
            if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
                if(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value==st.session_state.stanby_member1[number2]):
                    ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN,value="休憩")
                    break
    wb.save("data.xlsx")

    if mode=="[コートを空にする]":
        court_clear("Cコート",3)
    else:
        court_sorting("Cコート",3)

for number in range(4):
    if number==0:
        if ws_court.cell(row=number+2,column=COURT_WINLOSEA_COLUMN).value==1:
            df["勝敗A"]=True
        else:
            df["勝敗A"]=False
        df["メンバA"]=ws_court.cell(row=number+2,column=COURT_NAMEA_COLUMN).value
        if ws_court.cell(row=number+2,column=COURT_WINLOSEB_COLUMN).value==1:
            df["勝敗B"]=True
        else:
            df["勝敗B"]=False
        df["メンバB"]=ws_court.cell(row=number+2,column=COURT_NAMEB_COLUMN).value
        if ws_court.cell(row=number+2,column=COURT_WINLOSEC_COLUMN).value==1:
            df["勝敗C"]=True
        else:
            df["勝敗C"]=False
        df["メンバC"]=ws_court.cell(row=number+2,column=COURT_NAMEC_COLUMN).value
    else:
        if ws_court.cell(row=number+2,column=COURT_WINLOSEA_COLUMN).value==1:
            df_cache["勝敗A"]=True
        else:
            df_cache["勝敗A"]=False
        df_cache["メンバA"]=ws_court.cell(row=number+2,column=COURT_NAMEA_COLUMN).value
        if ws_court.cell(row=number+2,column=COURT_WINLOSEB_COLUMN).value==1:
            df_cache["勝敗B"]=True
        else:
            df_cache["勝敗B"]=False
        df_cache["メンバB"]=ws_court.cell(row=number+2,column=COURT_NAMEB_COLUMN).value
        if ws_court.cell(row=number+2,column=COURT_WINLOSEC_COLUMN).value==1:
            df_cache["勝敗C"]=True
        else:
            df_cache["勝敗C"]=False
        df_cache["メンバC"]=ws_court.cell(row=number+2,column=COURT_NAMEC_COLUMN).value

        df=pd.concat([df,df_cache],ignore_index=True)

edited_df = st.data_editor(df,use_container_width=True,hide_index=True)

#最新の値を変数に一時保存
for number in range(4):
    st.session_state.winloseA[number]=edited_df["勝敗A"][number]    #Aコート
    st.session_state.winloseB[number]=edited_df["勝敗B"][number]    #Bコート
    st.session_state.winloseC[number]=edited_df["勝敗C"][number]    #Cコート

##########################################################################################################
#StandByとBreakのインターフェイス
##########################################################################################################
stanby_member_original[0]["items"].clear()
stanby_member_original[1]["items"].clear()

#変数を定義/初期化(30名まで)
if 'stanby_member0' not in st.session_state:
    st.session_state.stanby_member0=["","","","","","","","","","","","","","","","","","","","","","","","","","","","","",""]
    st.session_state.stanby_member0_num=0
    st.session_state.stanby_member1=["","","","","","","","","","","","","","","","","","","","","","","","","","","","","",""]
    st.session_state.stanby_member1_num=0

for number in range(30):
    if ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value!=None:
        if(ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN).value=="待機"):
            stanby_member_original[0]["items"].append(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value)
        if(ws_member.cell(row=(number+2),column=MEMBER_STATUS_COLUMN).value=="休憩"):
            stanby_member_original[1]["items"].append(ws_member.cell(row=(number+2),column=MEMBER_NAME_COLUMN).value)
    else:
        break

st.session_state["sorted_data"]=sort_items(stanby_member_original, multi_containers=True,custom_style=custom_style)

#最新の値を変数に一時保存
st.session_state.stanby_member0_num=len(st.session_state["sorted_data"][0]["items"])
st.session_state.stanby_member1_num=len(st.session_state["sorted_data"][1]["items"])

for number in range(30):
    if st.session_state.stanby_member0_num<=number:
        st.session_state.stanby_member0[number]=""
    else:
        st.session_state.stanby_member0[number]=st.session_state["sorted_data"][0]["items"][number]

for number in range(30):
    if st.session_state.stanby_member1_num<=number:
        st.session_state.stanby_member1[number]=""
    else:
        st.session_state.stanby_member1[number]=st.session_state["sorted_data"][1]["items"][number]

##########################################################################################################
#履歴
##########################################################################################################

if st.button("履歴の削除"):
    for row in ws_history:
        for cell in row :
            cell.value = None
    ws_history.cell(row=1,column=HISTORY_NAME1_COLUMN,value="名前1")
    ws_history.cell(row=1,column=HISTORY_NAME2_COLUMN,value="名前2")
    ws_history.cell(row=1,column=HISTORY_NAME3_COLUMN,value="名前3")
    ws_history.cell(row=1,column=HISTORY_NAME4_COLUMN,value="名前4")
    ws_history.cell(row=1,column=HISTORY_COURT_COLUMN,value="コート")
    wb.save("data.xlsx")

df = pd.read_excel(r'data.xlsx',sheet_name='history')

st.table(df)
